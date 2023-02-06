import csv
import io

from openpyxl import load_workbook


class XlsxFile:

    def __init__(self, workbook):
        self.workbook = workbook


class XlsxReader:

    @staticmethod
    def validate_headers(headers_row, identifiers):
        valid = True if headers_row.get('url') and headers_row.get('value') and identifiers else False
        if not valid:
            raise Exception('Nombres de columnas incorrectos')

    @classmethod
    def get_headers(cls, worksheet):
        # Define data structure
        headers_row = {
            'container': '',
            'name': '',
            'url': '',
            'value': '',
            'notes': '',
        }
        identifiers = []
        # Iterate over first row looking for headers dict keys
        for row in worksheet.iter_rows(1, 1, 1):
            for cell in row:
                cell_value = cell.value if cell.value else None
                if cell_value and cell_value.lower() in headers_row.keys():
                    headers_row[cell_value.lower()] = cell.column
                elif cell_value:
                    identifiers.append({cell_value: cell.column})
                else:
                    break
        cls.validate_headers(headers_row, identifiers)
        return headers_row, identifiers

    @staticmethod
    def build_passwd_data(row, headers, identifiers):
        headers_dict_copy = headers.copy()
        identifiers_list_copy = identifiers.copy()
        # Update headers dict
        for field, column in headers_dict_copy.items():
            if column and isinstance(column, int):
                cell_value = row[column - 1].value
                headers_dict_copy[field] = cell_value if cell_value else None
            else:
                headers_dict_copy[field] = None
        # Update identifiers list
        filtered_identifiers = []
        for identifier in identifiers_list_copy:
            updated_identifier = {}
            for identifier_key, column in identifier.items():
                if column and isinstance(column, int):
                    cell_value = row[column - 1].value
                    if cell_value:
                        updated_identifier[identifier_key] = cell_value
                        filtered_identifiers.append(updated_identifier)
        headers_dict_copy['identifiers'] = filtered_identifiers
        return headers_dict_copy

    @staticmethod
    def validate_passwd_data(passwd_data):
        return passwd_data.get('value') and passwd_data.get('url')

    @classmethod
    def read_passwds_from_xlsx(cls, _bytes):
        workbook = load_workbook(_bytes)
        worksheet = workbook.active
        headers, identifiers = cls.get_headers(worksheet)
        passwds = []
        for row in worksheet.iter_rows(min_row=2, max_row=None, min_col=1, max_col=None):
            passwd_data = cls.build_passwd_data(row, headers, identifiers)
            if cls.validate_passwd_data(passwd_data):
                passwds.append(passwd_data)
        workbook.close()
        return passwds


class CsvReader:
    @classmethod
    def read_passwds_from_chrome_csv(cls, _bytes):
        file = _bytes.read().decode('utf-8')
        reader = csv.DictReader(io.StringIO(file))
        # Generate a list comprehension
        data = [line for line in reader]
        read_passwds = []
        for row in data:
            dict_passwd = dict(row)
            read_passwd = {
                "container": None,
                "name": dict_passwd["name"],
                "url": dict_passwd["url"],
                "value": dict_passwd["password"],
                "notes": None,
                "identifiers": [
                    {"username": dict_passwd["username"]},
                ]
            }
            read_passwds.append(read_passwd)

        return read_passwds


class FileReader(XlsxReader, CsvReader):
    pass
